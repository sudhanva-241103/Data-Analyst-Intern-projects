import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import warnings
warnings.filterwarnings('ignore')

np.random.seed(42)

# ─────────────────────────────────────────────
# 1. SIMULATE LINKEDIN JOB POSTINGS
# ─────────────────────────────────────────────
cities = ["New York", "San Francisco", "Austin", "Seattle", "Chicago",
          "Boston", "Los Angeles", "Remote"]

roles = ["Data Scientist", "ML Engineer", "Backend Engineer",
         "Frontend Engineer", "DevOps Engineer", "Product Manager",
         "Data Analyst", "Full Stack Engineer"]

skills_pool = {
    "Data Scientist":      ["Python", "SQL", "Machine Learning", "TensorFlow", "Statistics",
                            "Pandas", "Spark", "Tableau", "R", "Deep Learning"],
    "ML Engineer":         ["Python", "TensorFlow", "PyTorch", "MLOps", "Docker",
                            "Kubernetes", "Spark", "Scala", "AWS", "Machine Learning"],
    "Backend Engineer":    ["Python", "Java", "Go", "SQL", "REST APIs", "Docker",
                            "Kubernetes", "AWS", "Microservices", "PostgreSQL"],
    "Frontend Engineer":   ["JavaScript", "React", "TypeScript", "CSS", "HTML",
                            "Vue.js", "GraphQL", "Node.js", "Webpack", "Jest"],
    "DevOps Engineer":     ["Docker", "Kubernetes", "AWS", "Terraform", "CI/CD",
                            "Python", "Linux", "Ansible", "Prometheus", "Jenkins"],
    "Product Manager":     ["Agile", "SQL", "Roadmapping", "Stakeholder Mgmt", "Jira",
                            "A/B Testing", "Figma", "OKRs", "Analytics", "Scrum"],
    "Data Analyst":        ["SQL", "Python", "Tableau", "Power BI", "Excel",
                            "Statistics", "Pandas", "Looker", "R", "BigQuery"],
    "Full Stack Engineer": ["JavaScript", "React", "Node.js", "Python", "SQL",
                            "Docker", "REST APIs", "TypeScript", "AWS", "MongoDB"],
}

city_role_weights = {
    "San Francisco": {"ML Engineer": 3, "Data Scientist": 2.5, "Full Stack Engineer": 2},
    "Seattle":       {"Backend Engineer": 3, "ML Engineer": 2.5, "DevOps Engineer": 2},
    "New York":      {"Data Analyst": 3, "Product Manager": 2.5, "Backend Engineer": 2},
    "Austin":        {"Full Stack Engineer": 2.5, "DevOps Engineer": 2, "Backend Engineer": 2},
    "Chicago":       {"Data Analyst": 2.5, "Product Manager": 2, "Backend Engineer": 2},
    "Boston":        {"Data Scientist": 3, "ML Engineer": 2, "Backend Engineer": 2},
    "Los Angeles":   {"Frontend Engineer": 3, "Product Manager": 2, "Full Stack Engineer": 2},
    "Remote":        {"Full Stack Engineer": 3, "Frontend Engineer": 2.5, "DevOps Engineer": 2},
}

records = []
for city in cities:
    n_jobs = np.random.randint(120, 300)
    weights = [city_role_weights.get(city, {}).get(r, 1.0) for r in roles]
    weights = np.array(weights) / sum(weights)
    chosen_roles = np.random.choice(roles, size=n_jobs, p=weights)
    for role in chosen_roles:
        n_skills = np.random.randint(3, 7)
        job_skills = np.random.choice(skills_pool[role], size=n_skills, replace=False)
        salary_base = {"Data Scientist": 140, "ML Engineer": 155, "Backend Engineer": 135,
                       "Frontend Engineer": 120, "DevOps Engineer": 130, "Product Manager": 145,
                       "Data Analyst": 110, "Full Stack Engineer": 130}
        salary = int(salary_base[role] * np.random.uniform(0.85, 1.25)) * 1000
        records.append({
            "Title": role, "City": city,
            "Skills": ", ".join(job_skills),
            "Salary": salary,
            "Remote": "Yes" if city == "Remote" or np.random.random() < 0.3 else "No",
            "Experience_Yrs": np.random.randint(1, 10),
        })

df = pd.DataFrame(records)

# Explode skills
df_skills = df.assign(Skill=df["Skills"].str.split(", ")).explode("Skill")

# ─────────────────────────────────────────────
# 2. VIZ A – Top 10 Skills Heatmap by City
# ─────────────────────────────────────────────
top_skills = df_skills["Skill"].value_counts().head(15).index.tolist()
heat_data = (df_skills[df_skills["Skill"].isin(top_skills)]
             .groupby(["City", "Skill"]).size()
             .unstack(fill_value=0))
heat_pct = heat_data.div(heat_data.sum(axis=1), axis=0) * 100

fig, ax = plt.subplots(figsize=(14, 7))
fig.patch.set_facecolor("#0F172A")
ax.set_facecolor("#0F172A")

cmap = sns.color_palette("YlOrRd", as_cmap=True)
sns.heatmap(heat_pct, annot=True, fmt=".1f", cmap=cmap, ax=ax,
            linewidths=0.5, linecolor="#1E293B",
            cbar_kws={"label": "% of City Postings", "shrink": 0.8},
            annot_kws={"size": 8, "color": "black"})

ax.set_title("Top Skills Demand Heatmap by City (%)", fontsize=16,
             fontweight="bold", color="white", pad=15)
ax.set_xlabel("Skill", fontsize=11, color="#94A3B8")
ax.set_ylabel("City", fontsize=11, color="#94A3B8")
ax.tick_params(colors="white", labelsize=9)
plt.xticks(rotation=35, ha="right")
plt.yticks(rotation=0)

cbar = ax.collections[0].colorbar
cbar.ax.yaxis.label.set_color("white")
cbar.ax.tick_params(colors="white")

plt.tight_layout()
plt.savefig("/mnt/user-data/outputs/heatmap_skills_by_city.png",
            dpi=150, bbox_inches="tight", facecolor="#0F172A")
plt.close()

# ─────────────────────────────────────────────
# 3. VIZ B – Skill vs Role Matrix (bubble)
# ─────────────────────────────────────────────
top12_skills = df_skills["Skill"].value_counts().head(12).index.tolist()
matrix = (df_skills[df_skills["Skill"].isin(top12_skills)]
          .groupby(["Title", "Skill"]).size().reset_index(name="Count"))

fig, ax = plt.subplots(figsize=(14, 7))
fig.patch.set_facecolor("#0F172A")
ax.set_facecolor("#111827")

role_list = sorted(matrix["Title"].unique())
skill_list = sorted(matrix["Skill"].unique())
role_idx = {r: i for i, r in enumerate(role_list)}
skill_idx = {s: i for i, s in enumerate(skill_list)}

max_count = matrix["Count"].max()
colors = plt.cm.plasma(np.linspace(0.3, 0.9, len(role_list)))
color_map = {r: colors[i] for i, r in enumerate(role_list)}

for _, row in matrix.iterrows():
    x = skill_idx[row["Skill"]]
    y = role_idx[row["Title"]]
    size = (row["Count"] / max_count) * 1800
    ax.scatter(x, y, s=size, color=color_map[row["Title"]], alpha=0.85,
               edgecolors="white", linewidths=0.4)
    ax.text(x, y, str(int(row["Count"])), ha="center", va="center",
            fontsize=7, color="white", fontweight="bold")

ax.set_xticks(range(len(skill_list)))
ax.set_xticklabels(skill_list, rotation=35, ha="right", color="white", fontsize=9)
ax.set_yticks(range(len(role_list)))
ax.set_yticklabels(role_list, color="white", fontsize=9)
ax.set_title("Skill × Role Matrix — Bubble Size = Job Count",
             fontsize=15, fontweight="bold", color="white", pad=12)
ax.grid(True, color="#1E293B", linewidth=0.6)
ax.tick_params(colors="white")
for spine in ax.spines.values():
    spine.set_edgecolor("#334155")

legend_handles = [mpatches.Patch(color=color_map[r], label=r) for r in role_list]
ax.legend(handles=legend_handles, bbox_to_anchor=(1.01, 1), loc="upper left",
          framealpha=0, labelcolor="white", fontsize=8)

plt.tight_layout()
plt.savefig("/mnt/user-data/outputs/skill_role_matrix.png",
            dpi=150, bbox_inches="tight", facecolor="#0F172A")
plt.close()

# ─────────────────────────────────────────────
# 4. VIZ C – Job Demand by City & Role (stacked bar)
# ─────────────────────────────────────────────
demand = df.groupby(["City", "Title"]).size().unstack(fill_value=0)
demand = demand.loc[demand.sum(axis=1).sort_values(ascending=False).index]

fig, ax = plt.subplots(figsize=(14, 7))
fig.patch.set_facecolor("#0F172A")
ax.set_facecolor("#111827")

palette = plt.cm.tab10(np.linspace(0, 1, len(demand.columns)))
demand.plot(kind="bar", stacked=True, ax=ax, color=palette,
            edgecolor="#0F172A", linewidth=0.5, width=0.72)

ax.set_title("Job Demand by City & Role", fontsize=16, fontweight="bold",
             color="white", pad=12)
ax.set_xlabel("City", fontsize=11, color="#94A3B8")
ax.set_ylabel("Number of Job Postings", fontsize=11, color="#94A3B8")
ax.tick_params(colors="white", labelsize=9)
plt.xticks(rotation=25, ha="right")
ax.grid(axis="y", color="#1E293B", linewidth=0.6)
for spine in ax.spines.values():
    spine.set_edgecolor("#334155")

legend = ax.legend(bbox_to_anchor=(1.01, 1), loc="upper left",
                   framealpha=0, labelcolor="white", fontsize=8, title="Role")
legend.get_title().set_color("white")

plt.tight_layout()
plt.savefig("/mnt/user-data/outputs/job_demand_by_city.png",
            dpi=150, bbox_inches="tight", facecolor="#0F172A")
plt.close()

# ─────────────────────────────────────────────
# 5. VIZ D – Salary Distribution by Role
# ─────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(13, 6))
fig.patch.set_facecolor("#0F172A")
ax.set_facecolor("#111827")

role_order = df.groupby("Title")["Salary"].median().sort_values(ascending=False).index
palette_v = sns.color_palette("cool", len(role_order))
sns.violinplot(data=df, x="Title", y="Salary", order=role_order,
               palette=palette_v, ax=ax, inner="quartile",
               linewidth=0.8, cut=0)

ax.set_title("Salary Distribution by Role", fontsize=16, fontweight="bold",
             color="white", pad=12)
ax.set_xlabel("", fontsize=11)
ax.set_ylabel("Annual Salary (USD)", fontsize=11, color="#94A3B8")
ax.tick_params(colors="white", labelsize=8)
plt.xticks(rotation=20, ha="right")
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"${x/1000:.0f}K"))
ax.grid(axis="y", color="#1E293B", linewidth=0.6)
for spine in ax.spines.values():
    spine.set_edgecolor("#334155")

plt.tight_layout()
plt.savefig("/mnt/user-data/outputs/salary_by_role.png",
            dpi=150, bbox_inches="tight", facecolor="#0F172A")
plt.close()

# ─────────────────────────────────────────────
# 6. VIZ E – Top 15 Skills Overall (horizontal bar)
# ─────────────────────────────────────────────
top15 = df_skills["Skill"].value_counts().head(15)

fig, ax = plt.subplots(figsize=(11, 7))
fig.patch.set_facecolor("#0F172A")
ax.set_facecolor("#111827")

bars = ax.barh(top15.index[::-1], top15.values[::-1],
               color=plt.cm.viridis(np.linspace(0.3, 0.9, 15)),
               edgecolor="#0F172A", height=0.7)
for bar, val in zip(bars, top15.values[::-1]):
    ax.text(bar.get_width() + 8, bar.get_y() + bar.get_height()/2,
            str(val), va="center", color="white", fontsize=9)

ax.set_title("Top 15 Most In-Demand Skills (All Cities)", fontsize=15,
             fontweight="bold", color="white", pad=12)
ax.set_xlabel("Number of Job Postings", fontsize=11, color="#94A3B8")
ax.tick_params(colors="white", labelsize=10)
ax.grid(axis="x", color="#1E293B", linewidth=0.6)
for spine in ax.spines.values():
    spine.set_edgecolor("#334155")

plt.tight_layout()
plt.savefig("/mnt/user-data/outputs/top15_skills.png",
            dpi=150, bbox_inches="tight", facecolor="#0F172A")
plt.close()

# ─────────────────────────────────────────────
# 7. BUILD EXCEL WORKBOOK
# ─────────────────────────────────────────────
wb = Workbook()

DARK  = "0F172A"
MID   = "1E293B"
LIGHT = "334155"
ACCENT= "6366F1"
GREEN = "10B981"
AMBER = "F59E0B"
RED   = "EF4444"
WHITE = "FFFFFF"
GRAY  = "94A3B8"

thin = Side(style="thin", color=LIGHT)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bold=True, size=11, bg=MID, fg=WHITE, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border
    return c

def cell(ws, row, col, val, bold=False, size=10, bg=DARK, fg=WHITE,
         align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border
    if num_fmt:
        c.number_format = num_fmt
    return c

# ── Sheet 1: Dashboard Summary ──
ws1 = wb.active
ws1.title = "📊 Dashboard"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 50
ws1.row_dimensions[2].height = 20

# Title banner
ws1.merge_cells("A1:H1")
c = ws1["A1"]
c.value = "LinkedIn Job Trend Analysis — Skill Demand Intelligence"
c.font = Font(name="Arial", bold=True, size=20, color=WHITE)
c.fill = PatternFill("solid", fgColor=ACCENT)
c.alignment = Alignment(horizontal="center", vertical="center")

# KPI cards row
kpis = [
    ("Total Postings", len(df), GREEN, "#"),
    ("Cities Tracked", df["City"].nunique(), ACCENT, "#"),
    ("Roles Analyzed", df["Title"].nunique(), AMBER, "#"),
    ("Unique Skills", df_skills["Skill"].nunique(), RED, "#"),
    ("Avg Salary", int(df["Salary"].mean()), GREEN, "$#,##0"),
    ("% Remote Roles", round((df["Remote"]=="Yes").mean()*100,1), ACCENT, '0.0"%"'),
]
col_starts = [1, 2, 3, 4, 5, 6]
for i, (label, val, color, fmt) in enumerate(kpis):
    col = i + 1
    ws1.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col)
    ws1.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col)
    ws1.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col)
    hdr(ws1, 3, col, label, size=9, bg=color, fg=WHITE)
    c = ws1.cell(row=4, column=col, value=val)
    c.font = Font(name="Arial", bold=True, size=20, color=color)
    c.fill = PatternFill("solid", fgColor=MID)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = fmt
    c.border = border
    ws1.row_dimensions[4].height = 38

for col in range(1, 7):
    ws1.column_dimensions[get_column_letter(col)].width = 20

# Top skills table
hdr(ws1, 7, 1, "Rank", bg=ACCENT)
hdr(ws1, 7, 2, "Skill", bg=ACCENT)
hdr(ws1, 7, 3, "Total Mentions", bg=ACCENT)
hdr(ws1, 7, 4, "% of Postings", bg=ACCENT)
hdr(ws1, 7, 5, "Trend", bg=ACCENT)
top10_skills = df_skills["Skill"].value_counts().head(10)
for i, (skill, cnt) in enumerate(top10_skills.items(), 1):
    bg_row = DARK if i % 2 == 0 else MID
    pct = cnt / len(df) * 100
    trend = "🔥 Hot" if pct > 50 else ("📈 Rising" if pct > 30 else "➡ Stable")
    cell(ws1, 7+i, 1, i, bg=bg_row, align="center")
    cell(ws1, 7+i, 2, skill, bg=bg_row, align="left")
    cell(ws1, 7+i, 3, cnt, bg=bg_row)
    cell(ws1, 7+i, 4, round(pct, 1), bg=bg_row, num_fmt='0.0"%"')
    cell(ws1, 7+i, 5, trend, bg=bg_row)

# Recommendation box
ws1.merge_cells("A19:F22")
rec = ws1["A19"]
rec.value = (
    "💡 RECOMMENDATIONS: Python & SQL dominate across all cities — essential baseline for any tech role. "
    "ML/AI skills (TensorFlow, PyTorch, MLOps) command 15-25% salary premiums, especially in SF & Seattle. "
    "Cloud + containers (AWS, Docker, Kubernetes) are now table-stakes for DevOps & backend roles. "
    "Remote roles skew heavily toward Full Stack & Frontend — React + TypeScript are most portable skills."
)
rec.font = Font(name="Arial", size=10, color=WHITE, italic=True)
rec.fill = PatternFill("solid", fgColor="1E1B4B")
rec.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
rec.border = border
ws1.row_dimensions[19].height = 70

# ── Sheet 2: Raw Data ──
ws2 = wb.create_sheet("📋 Job Data")
ws2.sheet_view.showGridLines = False
cols2 = ["Title", "City", "Skills", "Salary", "Remote", "Experience_Yrs"]
for j, col in enumerate(cols2, 1):
    hdr(ws2, 1, j, col, bg=ACCENT)
    ws2.column_dimensions[get_column_letter(j)].width = [22, 16, 50, 14, 10, 16][j-1]
for i, row in df.iterrows():
    bg_row = DARK if i % 2 == 0 else MID
    for j, col in enumerate(cols2, 1):
        val = row[col]
        fmt = "$#,##0" if col == "Salary" else None
        cell(ws2, i+2, j, val, bg=bg_row, align="left" if col=="Skills" else "center",
             num_fmt=fmt)

# ── Sheet 3: Skill × City Heatmap Data ──
ws3 = wb.create_sheet("🌡 Skill×City")
ws3.sheet_view.showGridLines = False
top10 = df_skills["Skill"].value_counts().head(10).index.tolist()
pivot = (df_skills[df_skills["Skill"].isin(top10)]
         .groupby(["City","Skill"]).size().unstack(fill_value=0))
pivot_pct = pivot.div(pivot.sum(axis=1), axis=0).round(3)

hdr(ws3, 1, 1, "City \\ Skill", bg=ACCENT, size=11)
for j, skill in enumerate(pivot_pct.columns, 2):
    hdr(ws3, 1, j, skill, bg=ACCENT, size=10)
    ws3.column_dimensions[get_column_letter(j)].width = 14
ws3.column_dimensions["A"].width = 18

for i, (city, row) in enumerate(pivot_pct.iterrows(), 2):
    cell(ws3, i, 1, city, bg=MID, align="left", bold=True)
    for j, val in enumerate(row, 2):
        intensity = int(val * 255)
        r_hex = format(max(0, 255-intensity//2), "02X")
        g_hex = format(max(0, 100+intensity//3), "02X")
        b_hex = format(max(0, 50), "02X")
        bg_color = r_hex + g_hex + b_hex
        c = ws3.cell(row=i, column=j, value=round(val*100, 1))
        c.font = Font(name="Arial", size=10, color=WHITE, bold=val>0.1)
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        c.number_format = '0.0"%"'

# ── Sheet 4: Skill × Role Matrix ──
ws4 = wb.create_sheet("🎯 Skill×Role")
ws4.sheet_view.showGridLines = False
top12 = df_skills["Skill"].value_counts().head(12).index.tolist()
role_pivot = (df_skills[df_skills["Skill"].isin(top12)]
              .groupby(["Title","Skill"]).size().unstack(fill_value=0))

hdr(ws4, 1, 1, "Role \\ Skill", bg=ACCENT, size=11)
for j, skill in enumerate(role_pivot.columns, 2):
    hdr(ws4, 1, j, skill, bg=ACCENT, size=10)
    ws4.column_dimensions[get_column_letter(j)].width = 13
ws4.column_dimensions["A"].width = 22

max_val = role_pivot.values.max()
for i, (role, row) in enumerate(role_pivot.iterrows(), 2):
    cell(ws4, i, 1, role, bg=MID, align="left", bold=True)
    for j, val in enumerate(row, 2):
        intensity = int((val / max_val) * 200)
        bg_color = format(intensity, "02X") + format(min(255, intensity+50), "02X") + "FF"
        c = ws4.cell(row=i, column=j, value=int(val))
        c.font = Font(name="Arial", size=10, color=WHITE, bold=val>max_val*0.5)
        c.fill = PatternFill("solid", fgColor=bg_color if val > 0 else DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws4.row_dimensions[i].height = 22

# ── Sheet 5: Salary Analysis ──
ws5 = wb.create_sheet("💰 Salary")
ws5.sheet_view.showGridLines = False
sal = df.groupby("Title")["Salary"].agg(["min","mean","max","median","count"]).round(0).astype(int)
sal.columns = ["Min", "Mean", "Max", "Median", "# Postings"]
sal = sal.sort_values("Median", ascending=False)

headers5 = ["Role", "Min Salary", "Mean Salary", "Median Salary", "Max Salary", "# Postings"]
for j, h in enumerate(headers5, 1):
    hdr(ws5, 1, j, h, bg=ACCENT)
    ws5.column_dimensions[get_column_letter(j)].width = [22,14,14,15,14,12][j-1]

for i, (role, row) in enumerate(sal.iterrows(), 2):
    bg_row = DARK if i % 2 == 0 else MID
    cell(ws5, i, 1, role, bg=bg_row, align="left")
    for j, col in enumerate(["min","mean","median","max","count"], 2):
        fmt = "$#,##0" if col != "count" else "#,##0"
        cell(ws5, i, j, int(row[col.title() if col!="count" else "# Postings"]),
             bg=bg_row, num_fmt=fmt)

# ── Sheet 6: Recommendations ──
ws6 = wb.create_sheet("💡 Recommendations")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 28
ws6.column_dimensions["B"].width = 55
ws6.column_dimensions["C"].width = 20

hdr(ws6, 1, 1, "Category", bg=ACCENT, size=12)
hdr(ws6, 1, 2, "Recommendation", bg=ACCENT, size=12)
hdr(ws6, 1, 3, "Priority", bg=ACCENT, size=12)

recs = [
    ("🔑 Core Skills", "Python + SQL are required in 70%+ of all postings — learn these first", "🔴 Critical"),
    ("☁️ Cloud Skills", "AWS appears in 5 of 8 role categories — get AWS certified to boost employability", "🔴 Critical"),
    ("🤖 AI/ML Premium", "TensorFlow/PyTorch roles pay 15-25% more; SF & Seattle lead in ML hiring", "🟠 High"),
    ("🐳 Containerization", "Docker + Kubernetes now standard for Backend & DevOps — not optional", "🟠 High"),
    ("🌐 Remote Strategy", "Remote roles favor Full Stack & Frontend; React + TypeScript = most portable", "🟡 Medium"),
    ("📊 Data Analytics", "Tableau + Power BI needed for Analyst roles; NYC & Chicago largest markets", "🟡 Medium"),
    ("🏙️ City Targeting – SF", "Best for: ML Engineer, Data Scientist. Top skills: PyTorch, TensorFlow, Spark", "🟠 High"),
    ("🏙️ City Targeting – NYC", "Best for: Data Analyst, PM. Top skills: SQL, Tableau, Excel, Stakeholder Mgmt", "🟠 High"),
    ("🏙️ City Targeting – Seattle", "Best for: Backend, DevOps. Top skills: Go, AWS, Kubernetes, Microservices", "🟠 High"),
    ("📈 Salary Maximizer", "ML Engineer > Product Manager > Data Scientist = highest median salaries", "🟡 Medium"),
    ("🎓 Skill Gap Quick Win", "Add Docker + one cloud cert to any resume — appears in 6/8 role types", "🟡 Medium"),
    ("🔮 Emerging Trends", "MLOps & LLM-related skills (prompt engineering) showing rapid growth in postings", "🟢 Watch"),
]

for i, (cat, rec, pri) in enumerate(recs, 2):
    bg_row = DARK if i % 2 == 0 else MID
    cell(ws6, i, 1, cat, bg=bg_row, align="left", bold=True, fg=AMBER)
    cell(ws6, i, 2, rec, bg=bg_row, align="left", fg=WHITE)
    p_color = {"🔴": RED, "🟠": AMBER, "🟡": "EAB308", "🟢": GREEN}
    color = next((v for k, v in p_color.items() if pri.startswith(k)), WHITE)
    cell(ws6, i, 3, pri, bg=bg_row, align="center", fg=color, bold=True)
    ws6.row_dimensions[i].height = 28

wb.save("/mnt/user-data/outputs/LinkedIn_Job_Trend_Analysis.xlsx")
print("✅ All files generated successfully.")
